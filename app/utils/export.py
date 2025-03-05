import pandas as pd
from app.models import Fire, AuditLog, User
from flask import current_app
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

class FireExporter:
    @staticmethod
    def to_excel(fires, filename=None, start_date=None, end_date=None):
        if filename is None:
            filename = f'fires_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Фильтрация по датам
        if start_date:
            fires = [f for f in fires if f.date_reported >= start_date]
        if end_date:
            fires = [f for f in fires if f.date_reported <= end_date]
            
        data = [{
            'Дата пожара': fire.date_reported.strftime('%d.%м.%Y'),
            'Регион': fire.region,
            'КГУ/ООПТ': fire.kgu_oopt,
            'Филиал': fire.branch,
            'Лесничество': fire.forestry,
            'Квартал': fire.quarter,
            'Выдел': fire.vydel,
            'Площадь пожара': fire.area_total,
            'Площадь лесная': fire.area_forest,
            'Площадь лесная лесопокрытая': fire.area_covered,
            'Площадь верховой': fire.area_top,
            'Площадь не лесная': fire.area_non_forest,
            'Кол-во людей': fire.total_people,
            'Кол-во техники': fire.total_vehicles,
            'Ущерб (тенге)': fire.damage,
            'Затраты на тушение': fire.suppression_cost
        } for fire in fires]
        
        df = pd.DataFrame(data)
        export_path = os.path.join(current_app.config['EXPORT_PATH'], filename)
        
        # Сохранение с форматированием
        writer = pd.ExcelWriter(export_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Пожары')
        
        # Форматирование
        workbook = writer.book
        worksheet = writer.sheets['Пожары']
        
        # Заголовки
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        header_font = Font(bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
        writer.save()
        return filename

class DataExporter:
    @staticmethod
    def to_csv(start_date=None, end_date=None, region=None):
        query = Fire.query
        
        if start_date:
            query = query.filter(Fire.date_reported >= start_date)
        if end_date:
            query = query.filter(Fire.date_reported <= end_date)
        if region:
            query = query.filter(Fire.region == region)
            
        fires = query.all()
        
        data = [{
            'Дата пожара': fire.date_reported.strftime('%d.%м.%Y'),
            'Регион': fire.region,
            'КГУ/ООПТ': fire.kgu_oopt,
            'Филиал': fire.branch,
            'Лесничество': fire.forestry,
            'Квартал': fire.quarter,
            'Выдел': fire.vydel,
            'Площадь пожара': fire.area_total,
            'Площадь лесная': fire.area_forest,
            'Площадь лесная лесопокрытая': fire.area_covered,
            'Площадь верховой': fire.area_top,
            'Площадь не лесная': fire.area_non_forest,
            'Лесная охрана (люди)': fire.forest_guard_people,
            'Лесная охрана (техника)': fire.forest_guard_vehicles,
            'АПС (люди)': fire.aps_people,
            'АПС (техника)': fire.aps_vehicles,
            'АПС (авиация)': fire.aps_aircraft,
            'МЧС (люди)': fire.emergency_people,
            'МЧС (техника)': fire.emergency_vehicles,
            'МЧС (авиация)': fire.emergency_aircraft,
            'Ущерб (тенге)': fire.damage,
            'Затраты на тушение': fire.suppression_cost
        } for fire in fires]
        
        df = pd.DataFrame(data)
        filename = f'fires_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        export_path = os.path.join(current_app.config['EXPORT_PATH'], filename)
        
        df.to_csv(export_path, index=False, encoding='utf-8-sig')
        return filename

def export_to_excel():
    # Not implemented
    pass

def export_to_pdf():
    # Not implemented
    pass

def export_to_csv(start_date=None, end_date=None):
    query = Fire.query
    if start_date:
        query = query.filter(Fire.date_reported >= start_date)
    if end_date:
        query = query.filter(Fire.date_reported <= end_date)

def export_audit_logs(start_date=None, end_date=None):
    query = AuditLog.query.join(User)
    
    if start_date:
        query = query.filter(AuditLog.timestamp >= start_date)
    if end_date:
        query = query.filter(AuditLog.timestamp <= end_date)
        
    logs = query.all()
    
    data = [{
        'Время': log.timestamp.strftime('%d.%м.%Y %H:%M:%S'),
        'Пользователь': log.user.username,
        'Действие': log.action,
        'Таблица': log.table_name,
        'ID записи': log.record_id,
        'Изменения': log.changes
    } for log in logs]
    
    df = pd.DataFrame(data)
    filename = f'audit_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    df.to_csv(os.path.join(current_app.config['EXPORT_PATH'], filename), 
              index=False, encoding='utf-8-sig')
    return filename
